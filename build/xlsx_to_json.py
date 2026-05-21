#!/usr/bin/env python3
"""
xlsx_to_json.py — climatecaselab data pipeline (v1)

Reads a Sabin Center Global Trends XLSX snapshot and emits the five JSON files
that the static front-end consumes:

  src/data/countries.json   — per-jurisdiction summary (region, centroid, totals)
  src/data/flows.json       — directed country→country flows with citation counts
  src/data/cases.json       — per-source-case records with the cited cases attached
  src/data/catalytic.json   — top-5 source cases (data-driven, with exclusions)
  src/data/manifest.json    — snapshot label, hash, counts

It also writes build/qa_report.md summarising row drops and unknowns.

Usage:
  python build/xlsx_to_json.py \
      --xlsx refs/data/<file>.xlsx \
      --out  src/data/

The script intentionally lives **inside** climatecaselab/ and has no dependency
on the parent phdMutley pipeline.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
BUILD_DIR = ROOT / "build"
DEFAULT_XLSX = ROOT / "refs" / "data" / "7132427v1_Global_Trends_Data_v19.05.2026.xlsx"
DEFAULT_OUT = ROOT / "src" / "data"

# Columns we depend on, in display order.
REQUIRED_COLUMNS = [
    "Source Case",
    "Source Jurisdiction",
    "Source Region",
    "Citation Type",
    "Cited Case Name",
    "Case Law Origin",
    "Case Law Region",
    "Cited Year",
    "PDF URL",
    "Source Case URL",
    "Sabin Case ID (Cited)",
    "Verdict",
    "Bucket",
]

CONFIRMED_VERDICTS = {"CONFIRMED", "CORRECTED"}


# ---------------------------------------------------------------------------
# Loaders for the small static lookup files that travel with this script.
# ---------------------------------------------------------------------------


def _strip_comments(d: dict) -> dict:
    return {k: v for k, v in d.items() if not k.startswith("_")}


def load_aliases() -> dict[str, str]:
    raw = json.loads((BUILD_DIR / "country_aliases.json").read_text())
    return _strip_comments(raw)


def load_region_map() -> dict[str, str]:
    raw = json.loads((BUILD_DIR / "region_map.json").read_text())
    return _strip_comments(raw)


def load_centroids() -> dict[str, dict]:
    raw = json.loads((BUILD_DIR / "country_centroids.json").read_text())
    return _strip_comments(raw)


def load_exclusions() -> list[str]:
    raw = json.loads((BUILD_DIR / "catalytic_exclusions.json").read_text())
    return [s.lower() for s in raw.get("exclude_substrings", [])]


def load_catalytic_overrides() -> dict[str, Any]:
    """Curated metadata that the XLSX does not carry (preferred year, holdings slug).
    Color tokens are NOT keyed by case identity — every rank slot gets its own
    color via `--case-{rank}` in the consumer code (see spec §7.2)."""
    return {
        "case_urgenda": {
            "preferred_year": 2015,
            "preferred_name": "Urgenda Foundation v. State of the Netherlands",
        },
        "case_neubauer": {
            "preferred_year": 2021,
            "preferred_name": "Neubauer, et al. v. Germany",
        },
        "case_mass_v_epa": {
            "preferred_year": 2007,
            "preferred_name": "Massachusetts v. EPA",
        },
        "case_leghari": {
            "preferred_year": 2015,
            "preferred_name": "Leghari v. Federation of Pakistan",
        },
        "case_notre_affaire": {
            "preferred_year": 2021,
            "preferred_name": "Notre Affaire à Tous v. France",
        },
        "case_shell": {
            "preferred_year": 2021,
            "preferred_name": "Shell PLC v. Kingdom of the Netherlands",
        },
        "case_klimaseniorinnen": {
            "preferred_year": 2024,
            "preferred_name": "Verein KlimaSeniorinnen Schweiz and Others v. Switzerland",
        },
    }


CASE_SLUG_PATTERNS = [
    (re.compile(r"urgenda", re.I), "case_urgenda"),
    (re.compile(r"neubauer", re.I), "case_neubauer"),
    (re.compile(r"massachusetts\s+v\.?\s+epa", re.I), "case_mass_v_epa"),
    (re.compile(r"leghari", re.I), "case_leghari"),
    (re.compile(r"notre\s+affaire", re.I), "case_notre_affaire"),
    (re.compile(r"shell\s+plc", re.I), "case_shell"),
    (re.compile(r"klimaseniorinnen", re.I), "case_klimaseniorinnen"),
]


def slug_for_case(name: str) -> str | None:
    for pat, slug in CASE_SLUG_PATTERNS:
        if pat.search(name):
            return slug
    return None


# ---------------------------------------------------------------------------
# Data structures.
# ---------------------------------------------------------------------------


@dataclass
class Row:
    source_case: str
    source_jurisdiction: str
    source_region: str
    citation_type: str
    cited_case_name: str
    cited_origin: str
    cited_region: str
    cited_year: int | None
    pdf_url: str | None
    source_case_url: str | None
    sabin_id_cited: str | None
    verdict: str
    bucket: str


@dataclass
class QA:
    rows_total: int = 0
    rows_kept: int = 0
    rows_dropped: list[tuple[int, str]] = field(default_factory=list)
    unknown_countries: Counter = field(default_factory=Counter)
    region_conflicts: dict[str, set] = field(default_factory=lambda: defaultdict(set))
    odd_years: list[tuple[int, Any]] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Normalization helpers.
# ---------------------------------------------------------------------------


def canon_country(raw: str | None, aliases: dict[str, str]) -> str | None:
    if not raw or not str(raw).strip():
        return None
    s = str(raw).strip()
    return aliases.get(s, s)


def normalize_case_name(raw: str) -> str:
    """Strip trailing ' (YYYY)' and collapse whitespace so the same case under
    different spellings merges. Conservative — does not lowercase."""
    if not raw:
        return ""
    s = re.sub(r"\s*\(\d{4}\)\s*$", "", str(raw).strip())
    return re.sub(r"\s+", " ", s)


def safe_int(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        n = int(v)
        return n if 1900 <= n <= 2100 else None
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Core pipeline.
# ---------------------------------------------------------------------------


def read_rows(xlsx_path: Path, qa: QA) -> list[Row]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Planilha1"]
    headers = [c.value for c in ws[1]]
    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise SystemExit(f"❌ XLSX is missing required columns: {missing}")
    idx = {h: i for i, h in enumerate(headers)}

    rows: list[Row] = []
    for rn, vals in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        qa.rows_total += 1
        if not vals[idx["Source Case"]]:
            qa.rows_dropped.append((rn, "no Source Case"))
            continue
        verdict = (vals[idx["Verdict"]] or "").strip().upper()
        if verdict and verdict not in CONFIRMED_VERDICTS:
            qa.rows_dropped.append((rn, f"verdict={verdict!r}"))
            continue
        cited_name = vals[idx["Cited Case Name"]]
        if not cited_name:
            qa.rows_dropped.append((rn, "no Cited Case Name"))
            continue
        row = Row(
            source_case=str(vals[idx["Source Case"]]).strip(),
            source_jurisdiction=str(vals[idx["Source Jurisdiction"]] or "").strip(),
            source_region=str(vals[idx["Source Region"]] or "").strip(),
            citation_type=str(vals[idx["Citation Type"]] or "").strip(),
            cited_case_name=normalize_case_name(cited_name),
            cited_origin=str(vals[idx["Case Law Origin"]] or "").strip(),
            cited_region=str(vals[idx["Case Law Region"]] or "").strip(),
            cited_year=safe_int(vals[idx["Cited Year"]]),
            pdf_url=(vals[idx["PDF URL"]] or "") or None,
            source_case_url=(vals[idx["Source Case URL"]] or "") or None,
            sabin_id_cited=(vals[idx["Sabin Case ID (Cited)"]] or "") or None,
            verdict=verdict,
            bucket=(vals[idx["Bucket"]] or "") or "",
        )
        rows.append(row)
        qa.rows_kept += 1
    return rows


def build_countries(
    rows: list[Row],
    aliases: dict[str, str],
    region_map: dict[str, str],
    centroids: dict[str, dict],
    qa: QA,
) -> list[dict]:
    seen: dict[str, dict] = {}
    incoming = Counter()
    outgoing = Counter()
    for r in rows:
        src = canon_country(r.source_jurisdiction, aliases)
        tgt = canon_country(r.cited_origin, aliases)
        if src:
            outgoing[src] += 1
        if tgt:
            incoming[tgt] += 1

    all_names = set(incoming) | set(outgoing)
    for name in sorted(all_names):
        region = region_map.get(name)
        if not region:
            qa.unknown_countries[name] += 1
            region = "International"
        ctr = centroids.get(name)
        if not ctr:
            qa.unknown_countries[name] += 1
            continue
        seen[name] = {
            "name": name,
            "iso3": ctr.get("iso3"),
            "region": region,
            "lat": ctr["lat"],
            "lon": ctr["lon"],
            "incoming_total": incoming[name],
            "outgoing_total": outgoing[name],
        }
    return list(seen.values())


def build_flows(
    rows: list[Row],
    aliases: dict[str, str],
    region_map: dict[str, str],
    centroids: dict[str, dict],
) -> list[dict]:
    bucket: Counter = Counter()
    for r in rows:
        src = canon_country(r.source_jurisdiction, aliases)
        tgt = canon_country(r.cited_origin, aliases)
        if not src or not tgt or src == tgt:
            continue
        if src not in centroids or tgt not in centroids:
            continue
        bucket[(src, tgt)] += 1

    flows = []
    for i, ((src, tgt), n) in enumerate(sorted(bucket.items()), start=1):
        flows.append(
            {
                "id": f"f_{i:04d}",
                "source": src,
                "target": tgt,
                "source_region": region_map.get(src, "International"),
                "target_region": region_map.get(tgt, "International"),
                "count": n,
            }
        )
    return flows


def build_cases(
    rows: list[Row],
    aliases: dict[str, str],
    region_map: dict[str, str],
) -> list[dict]:
    grouped: dict[str, dict] = {}
    seen_pairs: dict[tuple[str, str], dict] = {}

    for r in rows:
        if r.source_case not in grouped:
            grouped[r.source_case] = {
                "name": r.source_case,
                "country": canon_country(r.source_jurisdiction, aliases),
                "region": r.source_region or "International",
                "url": r.source_case_url,
                "cited": [],
                "_seen": set(),
            }
        rec = grouped[r.source_case]
        key = (r.cited_case_name, canon_country(r.cited_origin, aliases) or "")
        if key in rec["_seen"]:
            continue
        rec["_seen"].add(key)
        country = canon_country(r.cited_origin, aliases) or r.cited_origin
        cited_region = region_map.get(country, r.cited_region or "International")
        rec["cited"].append(
            {
                "name": r.cited_case_name,
                "url": (
                    f"https://climatecasechart.com/case/{r.sabin_id_cited}/"
                    if r.sabin_id_cited
                    else None
                ),
                "country": country,
                "region": cited_region,
                "year": r.cited_year,
            }
        )

    cases = []
    for i, (name, rec) in enumerate(sorted(grouped.items()), start=1):
        rec.pop("_seen", None)
        cases.append(
            {
                "id": f"c_{i:04d}",
                "source_case_name": rec["name"],
                "source_case_url": rec["url"],
                "source_case_country": rec["country"],
                "source_case_region": rec["region"],
                "cited_cases": sorted(rec["cited"], key=lambda x: (-((x.get("country") or "") != ""), x["name"])),
            }
        )
    return cases


def build_catalytic(
    rows: list[Row],
    aliases: dict[str, str],
    region_map: dict[str, str],
    exclusions: list[str],
    overrides: dict[str, dict],
    top_n: int = 5,
) -> list[dict]:
    """
    Count rows per cited case (after name normalization), filter by exclusions,
    take top N, and assemble the rich shape the front-end expects.
    """
    by_case: dict[str, list[Row]] = defaultdict(list)
    for r in rows:
        by_case[r.cited_case_name].append(r)

    def is_excluded(name: str) -> bool:
        low = name.lower()
        return any(sub in low for sub in exclusions)

    # Score = total row count.
    ranked = sorted(
        ((name, group) for name, group in by_case.items() if not is_excluded(name)),
        key=lambda x: (-len(x[1]), x[0]),
    )[:top_n]

    out = []
    for rank, (name, group) in enumerate(ranked, start=1):
        slug = slug_for_case(name) or f"case_top{rank}"
        override = overrides.get(slug, {})
        years_counter: Counter = Counter()
        jurisdictions: Counter = Counter()
        jurisdiction_regions: dict[str, str] = {}
        citing_cases: list[dict] = []
        seen_citing = set()
        yearly_by_region: dict[int, Counter] = defaultdict(Counter)
        canonical_origin = None

        for r in group:
            src = canon_country(r.source_jurisdiction, aliases) or "Unknown"
            src_region = region_map.get(src, r.source_region or "International")
            jurisdictions[src] += 1
            jurisdiction_regions[src] = src_region
            if r.source_case_url:
                key = (r.source_case, src)
                if key not in seen_citing:
                    seen_citing.add(key)
                    citing_cases.append(
                        {
                            "name": r.source_case,
                            "country": src,
                            "year": r.cited_year,
                            "url": r.source_case_url,
                            "region": src_region,
                        }
                    )
            if r.cited_year:
                bucket = "north" if src_region == "Global North" else (
                    "south" if src_region == "Global South" else "international"
                )
                yearly_by_region[r.cited_year][bucket] += 1
            if not canonical_origin:
                canonical_origin = canon_country(r.cited_origin, aliases)

        sabin_ids = [r.sabin_id_cited for r in group if r.sabin_id_cited]
        sabin_url = (
            f"https://climatecasechart.com/case/{sabin_ids[0]}/"
            if sabin_ids
            else None
        )

        origin_country = canonical_origin or "Unknown"
        origin_region = region_map.get(origin_country, "International")

        # Order jurisdictions desc by count, then alpha.
        jurisdictions_sorted = [
            {"name": j, "count": c, "region": jurisdiction_regions[j]}
            for j, c in sorted(jurisdictions.items(), key=lambda x: (-x[1], x[0]))
        ]
        years_sorted = sorted(yearly_by_region.keys())
        yearly_payload = [
            {
                "year": y,
                "north": yearly_by_region[y]["north"],
                "south": yearly_by_region[y]["south"],
                "international": yearly_by_region[y]["international"],
            }
            for y in years_sorted
        ]

        out.append(
            {
                "id": slug,
                "rank": rank,
                "name": override.get("preferred_name") or name,
                "origin_country": origin_country,
                "origin_region": origin_region,
                "year": override.get("preferred_year"),
                "sabin_url": sabin_url,
                "color_token": f"--case-{rank}",
                "total_citing_decisions": sum(jurisdictions.values()),
                "total_jurisdictions": len(jurisdictions),
                "jurisdictions": jurisdictions_sorted,
                "citing_cases": sorted(citing_cases, key=lambda x: (x["country"] or "", x["name"])),
                "holdings_slug": override.get("holdings_slug") or slug,
                "yearly_citations": yearly_payload,
            }
        )
    return out


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def write_json(path: Path, obj: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, indent=2, ensure_ascii=False, sort_keys=False) + "\n")


def write_qa_report(qa: QA, out_path: Path) -> None:
    lines = [
        f"# QA report — climatecaselab build ({datetime.now(timezone.utc).isoformat(timespec='seconds')})",
        "",
        f"- Total rows in XLSX: **{qa.rows_total}**",
        f"- Rows kept after verdict + completeness filters: **{qa.rows_kept}**",
        f"- Rows dropped: **{len(qa.rows_dropped)}**",
    ]
    if qa.rows_dropped:
        lines.append("")
        lines.append("## Dropped rows")
        for rn, reason in qa.rows_dropped[:50]:
            lines.append(f"- row {rn}: {reason}")
        if len(qa.rows_dropped) > 50:
            lines.append(f"- … {len(qa.rows_dropped) - 50} more")
    if qa.unknown_countries:
        lines.append("")
        lines.append("## Unknown countries (no region or no centroid)")
        for name, n in qa.unknown_countries.most_common():
            lines.append(f"- {name}: {n}")
    lines.append("")
    out_path.write_text("\n".join(lines) + "\n")


# ---------------------------------------------------------------------------
# Entry point.
# ---------------------------------------------------------------------------


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    p.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = p.parse_args()

    if not args.xlsx.exists():
        print(f"❌ XLSX not found: {args.xlsx}", file=sys.stderr)
        return 1

    aliases = load_aliases()
    region_map = load_region_map()
    centroids = load_centroids()
    exclusions = load_exclusions()
    overrides = load_catalytic_overrides()
    qa = QA()

    try:
        rel = args.xlsx.resolve().relative_to(ROOT)
        label = str(rel)
    except ValueError:
        label = str(args.xlsx)
    print(f"▸ Reading {label}")
    rows = read_rows(args.xlsx, qa)
    print(f"  kept {qa.rows_kept}/{qa.rows_total} rows")

    countries = build_countries(rows, aliases, region_map, centroids, qa)
    flows = build_flows(rows, aliases, region_map, centroids)
    cases = build_cases(rows, aliases, region_map)
    catalytic = build_catalytic(rows, aliases, region_map, exclusions, overrides)

    manifest = {
        "snapshot_label": "Sabin Global Trends — May 2026",
        "snapshot_date": "2026-05-19",
        "xlsx_source_hash": sha256(args.xlsx),
        "case_count": len(cases),
        "country_count": len(countries),
        "flow_count": len(flows),
        "total_citations": sum(f["count"] for f in flows),
        "catalytic_count": len(catalytic),
        "built_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
    }

    write_json(args.out / "countries.json", countries)
    write_json(args.out / "flows.json", flows)
    write_json(args.out / "cases.json", cases)
    write_json(args.out / "catalytic.json", catalytic)
    write_json(args.out / "manifest.json", manifest)
    write_qa_report(qa, BUILD_DIR / "qa_report.md")

    print(f"✓ Wrote {len(countries)} countries, {len(flows)} flows, {len(cases)} cases, {len(catalytic)} catalytic")
    print(f"  manifest: {manifest['total_citations']} total citations across {manifest['country_count']} jurisdictions")
    if qa.unknown_countries:
        print(f"⚠ {len(qa.unknown_countries)} unknown countries — see build/qa_report.md")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
